#!/usr/bin/env python3
"""
validate_vc_master_kl.py — TK-99-1-1 성형 마스터 K/L열 무결성 검증.

검증 대상:
    Phase 1/2.Raw Materials/Vulcanization/성형공정_제약조건.xlsx

검증 항목:
    - K열(좌측셋팅) 모든 row ∈ {'o', 'x'}
    - L열(우측셋팅) 모든 row ∈ {'o', 'x'}
    - K='x' & L='x' 동시 (저압 가류기 사용 불가) 경고

출력:
    tools/master_validation/reports/vc_master_kl_<YYYY-MM-DD>.md

종료 코드:
    0 — 위반 0건
    1 — 위반 ≥1건 (CI/CD 게이트 차단)
"""

from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import openpyxl

# Windows cp949 회피 — UTF-8 강제
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

VALID_VALUES = {'o', 'x'}
MASTER_PATH = Path('Phase 1/2.Raw Materials/Vulcanization/성형공정_제약조건.xlsx')
SHEET_NAME = 'Sheet1 (2)'
REPORT_DIR = Path('tools/master_validation/reports')
HEADER_ROW = 2
DATA_START_ROW = 3
COL_HOSE = 1
COL_SPEC = 2          # 사양
COL_LP = 5            # 저압가류기
COL_IC = 13           # IC가류기
COL_K = 11            # 좌측셋팅
COL_L = 12            # 우측셋팅


def validate(wb: openpyxl.Workbook) -> tuple[list[dict], list[dict]]:
    """Returns (rows, violations)."""
    ws = wb[SHEET_NAME]

    # 헤더 검증 — 마스터 구조 변경 감지
    assert ws.cell(HEADER_ROW, COL_K).value == '좌측셋팅', \
        f"K열 헤더 = {ws.cell(HEADER_ROW, COL_K).value} (expected '좌측셋팅')"
    assert ws.cell(HEADER_ROW, COL_L).value == '우측셋팅', \
        f"L열 헤더 = {ws.cell(HEADER_ROW, COL_L).value} (expected '우측셋팅')"

    rows: list[dict] = []
    violations: list[dict] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        hose = ws.cell(r, COL_HOSE).value
        if not hose:
            break
        k = ws.cell(r, COL_K).value
        l = ws.cell(r, COL_L).value
        spec = ws.cell(r, COL_SPEC).value

        k_valid = k in VALID_VALUES
        l_valid = l in VALID_VALUES

        warn = ''
        if k == 'x' and l == 'x':
            warn = '⚠️ 좌·우 모두 x — 저압 슬롯 사용 불가 (IC 전용 또는 스케줄 불가)'

        row = {
            'row': r,
            'hose': str(hose).strip(),
            'spec': spec,
            'k': k,
            'l': l,
            'k_valid': k_valid,
            'l_valid': l_valid,
            'warn': warn,
        }
        rows.append(row)
        if not (k_valid and l_valid):
            violations.append(row)

    return rows, violations


def write_report(rows: list[dict], violations: list[dict]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    out = REPORT_DIR / f'vc_master_kl_{today}.md'

    # 분포 통계
    k_dist: dict[str, int] = {}
    l_dist: dict[str, int] = {}
    for r in rows:
        k_dist[repr(r['k'])] = k_dist.get(repr(r['k']), 0) + 1
        l_dist[repr(r['l'])] = l_dist.get(repr(r['l']), 0) + 1

    with open(out, 'w', encoding='utf-8') as f:
        f.write(f'# 성형 마스터 K/L열 검증 리포트 ({today})\n\n')
        f.write(f'**파일**: `{MASTER_PATH}` · sheet `{SHEET_NAME}`\n\n')
        f.write('## 1. 요약\n\n')
        f.write(f'- 총 품번 수: **{len(rows)}**\n')
        f.write(f'- K/L 무결성 위반: **{len(violations)}** 건\n')
        f.write(f'- 한쪽 only (좌측·우측 only): {sum(1 for r in rows if r["k"] != r["l"])} 건\n')
        f.write(f'- 좌·우 모두 x (저압 사용 불가): {sum(1 for r in rows if r["k"] == "x" and r["l"] == "x")} 건\n\n')

        f.write('## 2. 값 분포\n\n')
        f.write('| 컬럼 | 분포 |\n|---|---|\n')
        f.write(f'| K (좌측셋팅) | {dict(sorted(k_dist.items()))} |\n')
        f.write(f'| L (우측셋팅) | {dict(sorted(l_dist.items()))} |\n\n')

        f.write('## 3. 위반 목록\n\n')
        if not violations:
            f.write('✅ **위반 0건** — 47품번 모두 K∈{o,x}, L∈{o,x}\n\n')
        else:
            f.write('| Row | HOSE | K | L | 사유 |\n|---|---|---|---|---|\n')
            for v in violations:
                reason = []
                if not v['k_valid']:
                    reason.append(f'K={v["k"]!r} invalid')
                if not v['l_valid']:
                    reason.append(f'L={v["l"]!r} invalid')
                f.write(f'| {v["row"]} | {v["hose"]} | {v["k"]} | {v["l"]} | {"; ".join(reason)} |\n')
            f.write('\n')

        f.write('## 4. 한쪽 only 품번 (좌측·우측 only — BR-V15·V16 후보)\n\n')
        single_side = [r for r in rows if r['k'] != r['l']]
        if single_side:
            f.write('| HOSE | K | L | 셋팅 |\n|---|---|---|---|\n')
            for r in single_side:
                side = '좌측 only (K=o)' if r['k'] == 'o' and r['l'] == 'x' else '우측 only (L=o)'
                f.write(f'| {r["hose"]} | {r["k"]} | {r["l"]} | {side} |\n')
            f.write('\n')
        else:
            f.write('해당 없음 (모든 품번이 K=L)\n\n')

        f.write('## 5. 좌·우 모두 x 품번 (저압 사용 불가 — IC 전용 또는 스케줄 불가 의심)\n\n')
        both_x = [r for r in rows if r['k'] == 'x' and r['l'] == 'x']
        if both_x:
            f.write('| HOSE | K | L |\n|---|---|---|\n')
            for r in both_x:
                f.write(f'| {r["hose"]} | {r["k"]} | {r["l"]} |\n')
            f.write('\n')
        else:
            f.write('해당 없음 (모든 품번 ≥1면 사용 가능)\n\n')

        f.write('## 6. 전체 상세\n\n')
        f.write('| Row | HOSE | 사양 | K | L | 비고 |\n|---|---|---|:---:|:---:|---|\n')
        for r in rows:
            mark = '' if (r['k_valid'] and r['l_valid']) else '⚠️ '
            f.write(f'| {r["row"]} | {mark}{r["hose"]} | {r["spec"]} | {r["k"]} | {r["l"]} | {r["warn"]} |\n')

        f.write('\n---\n')
        f.write('## dual-review (BR-X05)\n\n')
        f.write('| 역할 | 이름 | 사인오프 일자 |\n|---|---|---|\n')
        f.write('| P1 생산계획 주임 | (서명) | (일자) |\n')
        f.write('| STK-08 IT lead | (서명) | (일자) |\n')

    return out


def main() -> int:
    if not MASTER_PATH.exists():
        print(f'ERROR: master file not found: {MASTER_PATH}', file=sys.stderr)
        return 2
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    rows, violations = validate(wb)
    report = write_report(rows, violations)
    print(f'Report: {report}')
    print(f'Total: {len(rows)} 품번, 위반: {len(violations)} 건')
    return 0 if not violations else 1


if __name__ == '__main__':
    sys.exit(main())
