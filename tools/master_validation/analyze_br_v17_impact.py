#!/usr/bin/env python3
"""
analyze_br_v17_impact.py — TK-99-2-3 BR-V17 영향 품번 운영 점검.

분석 항목:
    1. BR-V17 (규격<7) 영향 7품번 식별
    2. BR-V14·V15·V16 특수 제약 품번과 중첩 여부
    3. 수주 빈도 — 실 수주 데이터(`Phase 1/2.Raw Materials/Order/*.xlsx`) 기반 평균 빈도

출력:
    tools/master_validation/reports/br_v17_impact_<YYYY-MM-DD>.md
"""

from __future__ import annotations

import io
import sys
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EX_MASTER = Path('Phase 1/2.Raw Materials/Extrusion/압출공정_제약조건.xlsx')
ORDER_DIR = Path('Phase 1/2.Raw Materials/Order')
REPORT_DIR = Path('tools/master_validation/reports')

# BR-V14·V15·V16 특수 품번
SPECIAL_HOSES = {
    '28422-08HA0': 'BR-V14 (LP 단일 호기)',
    '28422-2M800': 'BR-V15 (LP 우측 only)',
    '28421-2M800': 'BR-V16 (LP 좌측 only)',
}


def load_ex_v17_impact() -> list[dict]:
    wb = openpyxl.load_workbook(EX_MASTER, data_only=True)
    ws = wb['Sheet1']
    rows = []
    for r in range(2, ws.max_row + 1):
        hose = ws.cell(r, 1).value
        if not hose:
            break
        spec = ws.cell(r, 2).value
        if isinstance(spec, (int, float)) and spec < 7:
            rows.append({
                'hose': str(hose).strip(),
                'spec': spec,
                'inner': ws.cell(r, 3).value,
                'thick': ws.cell(r, 4).value,
                'composite': ws.cell(r, 8).value,
            })
    return rows


def count_order_frequency(impact_hoses: list[str]) -> dict[str, int]:
    """수주 엑셀 파일 안의 hose_id 출현 빈도 (단순 카운트). NULL 또는 미발견 OK."""
    counter: Counter[str] = Counter()
    if not ORDER_DIR.exists():
        return dict(counter)

    target_set = set(impact_hoses)
    for xlsx in ORDER_DIR.glob('*.xlsx'):
        if xlsx.name.startswith('~$'):
            continue
        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
        except Exception as e:
            print(f'Skip {xlsx.name}: {e}', file=sys.stderr)
            continue
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        cell_clean = cell.strip()
                        if cell_clean in target_set:
                            counter[cell_clean] += 1
        wb.close()
    return dict(counter)


def write_report(impact: list[dict], freq: dict[str, int]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    out = REPORT_DIR / f'br_v17_impact_{today}.md'

    overlap_count = sum(1 for r in impact if r['hose'] in SPECIAL_HOSES)
    total_freq = sum(freq.values())

    with open(out, 'w', encoding='utf-8') as f:
        f.write(f'# BR-V17 영향 품번 운영 점검 ({today})\n\n')
        f.write('BR-V17 (PDD-02 §9): 규격<7 hose_id 는 1대 가류기당 동시 앵글 점유 ≤ 4.\n\n')

        f.write('## 1. 요약\n\n')
        f.write(f'- BR-V17 영향 품번: **{len(impact)}** (규격<7)\n')
        f.write(f'- BR-V14·V15·V16 (특수제약 3종) 중첩: **{overlap_count}**\n')
        f.write(f'- 수주 엑셀 총 출현 횟수: {total_freq}\n')
        f.write(f'- 수주 데이터 출처: `{ORDER_DIR}/*.xlsx`\n\n')

        f.write('## 2. 영향 품번 + 운영 점검\n\n')
        f.write('| Hose | 규격 | 내경 | 두께 | 합금형 | 특수제약 | 수주 출현 |\n')
        f.write('|---|---:|---:|---:|---:|---|---:|\n')
        for r in impact:
            overlap = SPECIAL_HOSES.get(r['hose'], '—')
            cnt = freq.get(r['hose'], 0)
            f.write(f'| {r["hose"]} | {r["spec"]} | {r["inner"]} | {r["thick"]} | {r["composite"]} | {overlap} | {cnt} |\n')
        f.write('\n')

        if overlap_count == 0:
            f.write('## 3. 중첩 분석\n\n')
            f.write('✅ **BR-V17 영향 품번 중 BR-V14·V15·V16 대상 없음** — 룰 cascade 위험 0.\n\n')
        else:
            f.write('## 3. 중첩 분석\n\n')
            f.write('⚠️ BR-V17 영향 품번이 BR-V14·V15·V16 과 중첩됨 — 룰 cascade 검증 필요.\n\n')
            f.write('| Hose | BR-V17 | 추가 BR |\n|---|---|---|\n')
            for r in impact:
                if r['hose'] in SPECIAL_HOSES:
                    f.write(f'| {r["hose"]} | ✓ | {SPECIAL_HOSES[r["hose"]]} |\n')
            f.write('\n')

        f.write('## 4. 수주 빈도 분석\n\n')
        if total_freq == 0:
            f.write('⚠️ 수주 엑셀에서 영향 품번 출현 0건 — 다음 중 하나:\n')
            f.write('1. 수주 엑셀 컬럼 구조가 hose_id 직접 포함하지 않음\n')
            f.write('2. 영향 품번이 실제로 수주되지 않음 (저빈도)\n')
            f.write('3. 수주 엑셀 hose_id 포맷 불일치 (공백·하이픈)\n\n')
            f.write('→ Sprint 1+ 수주 통합 (EP-01) 후 정밀 분석.\n\n')
        else:
            f.write(f'영향 품번 출현 빈도 (수주 엑셀 {len(list(ORDER_DIR.glob("*.xlsx")))}개 파일 기준):\n\n')
            f.write('| Hose | 출현 횟수 | 운영 priority |\n|---|---:|:---:|\n')
            for r in impact:
                cnt = freq.get(r['hose'], 0)
                pri = '🔴 High' if cnt >= 20 else ('🟡 Mid' if cnt >= 5 else '🟢 Low')
                f.write(f'| {r["hose"]} | {cnt} | {pri} |\n')
            f.write('\n')

        f.write('## 5. 권장 운영 조치\n\n')
        f.write(f'1. **Sprint 2 EP-21 ST-21-5** (규격<7 가류기당 ≤4 강제) 본격 구현 시 본 {len(impact)} 품번을 회귀 테스트 fixture 로 사용\n')
        f.write('2. **수주 빈도 ≥ 5건/월** 품번은 Grafana KPI 대시보드 별 panel 추가 검토\n')
        f.write(f'3. **VC_HOSE_RULE 시드 SQL** 작성 시 BR-V17 영향 {len(impact)} 품번 일괄 INSERT\n')
        f.write('4. **호싱불량 통계** cross-check 는 사내 품질부서 자료 입수 후 추가 분석 (현재 데이터 미보유)\n\n')

        f.write('\n---\n')
        f.write('## dual-review (BR-X05)\n\n')
        f.write('| 역할 | 이름 | 사인오프 일자 |\n|---|---|---|\n')
        f.write('| P1 생산계획 주임 | (서명) | (일자) |\n')
        f.write('| STK-08 IT lead | (서명) | (일자) |\n')

    return out


def main() -> int:
    impact = load_ex_v17_impact()
    if not impact:
        print('No spec<7 hoses found — nothing to analyze.')
        return 0
    freq = count_order_frequency([r['hose'] for r in impact])
    report = write_report(impact, freq)
    print(f'Report: {report}')
    print(f'영향 품번: {len(impact)} / 수주 출현 합계: {sum(freq.values())}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
