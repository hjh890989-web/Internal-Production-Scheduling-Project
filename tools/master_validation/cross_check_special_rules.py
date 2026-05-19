#!/usr/bin/env python3
"""
cross_check_special_rules.py — TK-99-1-2 품번별 특수 제약 cross-check.

대상 품번:
    - 28422-08HA0  (BR-V14: LP 단일 호기 셋팅, max_concurrent_slots=1)
    - 28422-2M800  (BR-V15: LP 우측 only, max_concurrent_slots=2)
    - 28421-2M800  (BR-V16: LP 좌측 only, max_concurrent_slots=2)

cross-check 항목:
    1. 성형 마스터 K/L 값 ↔ BR 명세
    2. 합금형 수량 + 가류기 정보
    3. VC_HOSE_RULE 시드 후 cross-reference (Sprint 1+ Flyway 후)

출력:
    tools/master_validation/reports/cross_check_special_rules_<YYYY-MM-DD>.md

종료 코드:
    0 — 모든 항목 명세 일치
    1 — 1건 이상 불일치
"""

from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

VC_MASTER = Path('Phase 1/2.Raw Materials/Vulcanization/성형공정_제약조건.xlsx')
SHEET_VC = 'Sheet1 (2)'
REPORT_DIR = Path('tools/master_validation/reports')

# 기대 명세 — BR-V14·V15·V16 / REQ-FUNC-VC-024·025·026 / 본 메모리 project_vc_constraint_v1_1
EXPECTED = {
    '28422-08HA0': {
        'br': 'BR-V14',
        'req': 'REQ-FUNC-VC-024',
        'description': 'LP 단일 호기 (LP-01) 단일 셋팅. max_concurrent_slots=1.',
        'k': 'o',          # 좌측 사용 가능 — 단, 동시 슬롯 ≤1 강제
        'l': 'o',          # 우측 사용 가능 — 단, 동시 슬롯 ≤1 강제
        'side_only': None, # 한쪽 only 아님
        'max_concurrent_slots': 1,
        'lp_only': True,   # IC 가류기 사용 불가
    },
    '28422-2M800': {
        'br': 'BR-V15',
        'req': 'REQ-FUNC-VC-025',
        'description': 'LP 우측 only. max_concurrent_slots=2.',
        'k': 'x',          # 좌측 사용 불가
        'l': 'o',          # 우측 only
        'side_only': 'right',
        'max_concurrent_slots': 2,
        'lp_only': False,
    },
    '28421-2M800': {
        'br': 'BR-V16',
        'req': 'REQ-FUNC-VC-026',
        'description': 'LP 좌측 only. max_concurrent_slots=2.',
        'k': 'o',          # 좌측 only
        'l': 'x',          # 우측 사용 불가
        'side_only': 'left',
        'max_concurrent_slots': 2,
        'lp_only': False,
    },
}

HEADER_ROW = 2
DATA_START_ROW = 3
COL_HOSE = 1
COL_SPEC = 2
COL_MOLD_CNT = 3      # 금형보유수량
COL_COMPOSITE = 4     # 합금형 (1 / 2 / 3 / 6)
COL_LP = 5            # 저압가류기
COL_LP_ANGLE = 6      # 저압앵글보유수량
COL_K = 11            # 좌측셋팅
COL_L = 12            # 우측셋팅
COL_IC = 13           # IC가류기
COL_IC_ANGLE = 14     # IC앵글보유수량


def find_actual(wb: openpyxl.Workbook, hose_id: str) -> dict | None:
    ws = wb[SHEET_VC]
    for r in range(DATA_START_ROW, ws.max_row + 1):
        h = ws.cell(r, COL_HOSE).value
        if not h:
            break
        if str(h).strip() == hose_id:
            return {
                'row': r,
                'hose': str(h).strip(),
                'spec': ws.cell(r, COL_SPEC).value,
                'mold_count': ws.cell(r, COL_MOLD_CNT).value,
                'composite': ws.cell(r, COL_COMPOSITE).value,
                'lp_machine': ws.cell(r, COL_LP).value,
                'lp_angle': ws.cell(r, COL_LP_ANGLE).value,
                'k': ws.cell(r, COL_K).value,
                'l': ws.cell(r, COL_L).value,
                'ic_machine': ws.cell(r, COL_IC).value,
                'ic_angle': ws.cell(r, COL_IC_ANGLE).value,
            }
    return None


def compare(expected: dict, actual: dict | None) -> list[str]:
    issues: list[str] = []
    if actual is None:
        return ['❌ HOSE 미발견 (마스터 엑셀에 없음)']
    if actual['k'] != expected['k']:
        issues.append(f"K(좌측) {actual['k']!r} ≠ expected {expected['k']!r}")
    if actual['l'] != expected['l']:
        issues.append(f"L(우측) {actual['l']!r} ≠ expected {expected['l']!r}")
    if expected['lp_only']:
        ic_val = actual['ic_machine']
        if ic_val and ic_val != 0:
            issues.append(f"IC가류기 {ic_val} ≠ 0 (LP only 위반)")
    return issues


def write_report(results: list[dict]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    out = REPORT_DIR / f'cross_check_special_rules_{today}.md'

    total_issues = sum(len(r['issues']) for r in results)

    with open(out, 'w', encoding='utf-8') as f:
        f.write(f'# 품번 특수 제약 cross-check 리포트 ({today})\n\n')
        f.write('대상: `28422-08HA0` (BR-V14) · `28422-2M800` (BR-V15) · `28421-2M800` (BR-V16)\n\n')

        f.write('## 1. 요약\n\n')
        f.write(f'- 검사 품번: 3 (BR-V14·V15·V16)\n')
        f.write(f'- 불일치 발견: **{total_issues}** 건\n\n')

        for r in results:
            f.write(f'## 2.{["", "1", "2", "3"][results.index(r) + 1]} {r["hose"]} — {r["expected"]["br"]}\n\n')
            f.write(f'**룰**: {r["expected"]["description"]}\n\n')
            f.write(f'**SRS**: {r["expected"]["req"]}\n\n')

            if r['actual'] is None:
                f.write('❌ 마스터에서 발견 못 함\n\n')
                continue

            f.write('### 기대 vs 실측\n\n')
            f.write('| 항목 | 기대 | 실측 |\n|---|---|---|\n')
            f.write(f'| K(좌측) | `{r["expected"]["k"]}` | `{r["actual"]["k"]}` |\n')
            f.write(f'| L(우측) | `{r["expected"]["l"]}` | `{r["actual"]["l"]}` |\n')
            f.write(f'| LP 가류기 | (any) | `{r["actual"]["lp_machine"]}` |\n')
            f.write(f'| LP 앵글수 | (any) | `{r["actual"]["lp_angle"]}` |\n')
            f.write(f'| IC 가류기 | {"0 (LP only)" if r["expected"]["lp_only"] else "(any)"} | `{r["actual"]["ic_machine"]}` |\n')
            f.write(f'| 합금형 | (any) | `{r["actual"]["composite"]}` |\n')
            f.write(f'| 사양 | (any) | `{r["actual"]["spec"]}` |\n\n')

            f.write('### cross-check 결과\n\n')
            if not r['issues']:
                f.write('✅ **모든 항목 명세 일치**\n\n')
            else:
                for issue in r['issues']:
                    f.write(f'- ⚠️ {issue}\n')
                f.write('\n')

            f.write(f'**VC_HOSE_RULE 시드 SQL 권장** (Sprint 1+ EP-21 Flyway):\n\n')
            f.write('```sql\n')
            f.write('INSERT INTO master.vc_hose_rule (hose_id, br_code, side_only, max_concurrent_slots, lp_only)\n')
            f.write(f'VALUES (\'{r["hose"]}\', \'{r["expected"]["br"]}\', '
                    f'{repr(r["expected"]["side_only"]) if r["expected"]["side_only"] else "NULL"}, '
                    f'{r["expected"]["max_concurrent_slots"]}, '
                    f'{str(r["expected"]["lp_only"]).upper()});\n')
            f.write('```\n\n')

        f.write('\n---\n')
        f.write('## dual-review (BR-X05)\n\n')
        f.write('| 역할 | 이름 | 사인오프 일자 |\n|---|---|---|\n')
        f.write('| P1 생산계획 주임 | (서명) | (일자) |\n')
        f.write('| STK-08 IT lead | (서명) | (일자) |\n')

    return out


def main() -> int:
    if not VC_MASTER.exists():
        print(f'ERROR: VC master not found: {VC_MASTER}', file=sys.stderr)
        return 2
    wb = openpyxl.load_workbook(VC_MASTER, data_only=True)
    results = []
    for hose_id, exp in EXPECTED.items():
        actual = find_actual(wb, hose_id)
        issues = compare(exp, actual)
        results.append({'hose': hose_id, 'expected': exp, 'actual': actual, 'issues': issues})

    report = write_report(results)
    total_issues = sum(len(r['issues']) for r in results)
    print(f'Report: {report}')
    print(f'cross-check: 3 품번, 불일치: {total_issues} 건')
    return 0 if total_issues == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
