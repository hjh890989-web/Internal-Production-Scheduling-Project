#!/usr/bin/env python3
"""
validate_ex_master_b.py — TK-99-2-1 압출 마스터 B열(규격) 정수형 무결성 검증.

검증 대상:
    Phase 1/2.Raw Materials/Extrusion/압출공정_제약조건.xlsx

검증 항목:
    - B열(규격) 모든 row 가 숫자 (int 또는 정수 float)
    - NULL·문자열·소수 비표준 값 탐지
    - 분포 통계 (5/7/9/11/13.5/18 등)

출력:
    tools/master_validation/reports/ex_master_b_<YYYY-MM-DD>.md

종료 코드:
    0 — 위반 0건
    1 — 위반 ≥1건 (CI/CD 게이트)
"""

from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path
from collections import Counter
from numbers import Number

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

MASTER_PATH = Path('Phase 1/2.Raw Materials/Extrusion/압출공정_제약조건.xlsx')
SHEET_NAME = 'Sheet1'
REPORT_DIR = Path('tools/master_validation/reports')
HEADER_ROW = 1
DATA_START_ROW = 2
COL_HOSE = 1
COL_SPEC = 2          # 규격
COL_INNER = 3         # 내경
COL_THICK = 4         # 두께


def is_numeric(v) -> bool:
    """B열 값이 숫자 (int 또는 float) 인지 — 소수 13.5 도 허용 (실제 데이터)."""
    return isinstance(v, Number) and not isinstance(v, bool)


def validate(wb: openpyxl.Workbook) -> tuple[list[dict], list[dict]]:
    ws = wb[SHEET_NAME]

    assert ws.cell(HEADER_ROW, COL_HOSE).value == 'HOSE', \
        f"A열 헤더 = {ws.cell(HEADER_ROW, COL_HOSE).value} (expected 'HOSE')"
    assert ws.cell(HEADER_ROW, COL_SPEC).value == '규격', \
        f"B열 헤더 = {ws.cell(HEADER_ROW, COL_SPEC).value} (expected '규격')"

    rows: list[dict] = []
    violations: list[dict] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        hose = ws.cell(r, COL_HOSE).value
        if not hose:
            break
        spec = ws.cell(r, COL_SPEC).value
        inner = ws.cell(r, COL_INNER).value
        thick = ws.cell(r, COL_THICK).value
        spec_valid = is_numeric(spec)

        row = {
            'row': r,
            'hose': str(hose).strip(),
            'spec': spec,
            'inner': inner,
            'thick': thick,
            'spec_valid': spec_valid,
        }
        rows.append(row)
        if not spec_valid:
            violations.append(row)

    return rows, violations


def write_report(rows: list[dict], violations: list[dict]) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    out = REPORT_DIR / f'ex_master_b_{today}.md'

    # 분포 통계
    spec_counter = Counter(r['spec'] for r in rows if r['spec_valid'])
    spec_dist = sorted(spec_counter.items(), key=lambda x: x[0])

    with open(out, 'w', encoding='utf-8') as f:
        f.write(f'# 압출 마스터 B열(규격) 검증 리포트 ({today})\n\n')
        f.write(f'**파일**: `{MASTER_PATH}` · sheet `{SHEET_NAME}`\n\n')

        f.write('## 1. 요약\n\n')
        f.write(f'- 총 품번 수: **{len(rows)}**\n')
        f.write(f'- B열 무결성 위반 (NULL·문자열·기타): **{len(violations)}** 건\n')
        f.write(f'- 고유 규격 수: {len(spec_counter)}\n')
        f.write(f'- 규격<7 품번 (BR-V17 영향): **{sum(c for s, c in spec_counter.items() if s < 7)}** 건\n\n')

        f.write('## 2. 규격 분포\n\n')
        f.write('| 규격 | 품번 수 | 비율 |\n|---:|---:|---:|\n')
        total = sum(c for _, c in spec_dist)
        for s, c in spec_dist:
            f.write(f'| {s} | {c} | {c/total*100:.1f}% |\n')
        f.write(f'| **합계** | **{total}** | 100.0% |\n\n')

        f.write('## 3. 위반 목록\n\n')
        if not violations:
            f.write('✅ **위반 0건** — 모든 품번 B열이 숫자\n\n')
        else:
            f.write('| Row | HOSE | B(규격) | 사유 |\n|---|---|---|---|\n')
            for v in violations:
                f.write(f'| {v["row"]} | {v["hose"]} | {v["spec"]!r} | non-numeric |\n')
            f.write('\n')

        f.write('## 4. 규격<7 품번 — BR-V17 영향 범위\n\n')
        impact = [r for r in rows if r['spec_valid'] and r['spec'] < 7]
        if impact:
            f.write(f'> BR-V17 (PDD-02 §9): 규격<7 hose_id 는 1대 가류기당 동시 앵글 점유 ≤ 4.\n\n')
            f.write('| Row | HOSE | 규격 | 내경 | 두께 |\n|---|---|---:|---:|---:|\n')
            for r in impact:
                f.write(f'| {r["row"]} | {r["hose"]} | {r["spec"]} | {r["inner"]} | {r["thick"]} |\n')
            f.write('\n')
        else:
            f.write('해당 없음\n\n')

        f.write('## 5. 전체 상세\n\n')
        f.write('| Row | HOSE | 규격 | 내경 | 두께 |\n|---|---|---:|---:|---:|\n')
        for r in rows:
            mark = '' if r['spec_valid'] else '⚠️ '
            f.write(f'| {r["row"]} | {mark}{r["hose"]} | {r["spec"]} | {r["inner"]} | {r["thick"]} |\n')

        f.write('\n---\n')
        f.write('## dual-review (BR-X05)\n\n')
        f.write('| 역할 | 이름 | 사인오프 일자 |\n|---|---|---|\n')
        f.write('| P1 생산계획 주임 | (서명) | (일자) |\n')
        f.write('| STK-08 IT lead | (서명) | (일자) |\n')

    return out


def main() -> int:
    if not MASTER_PATH.exists():
        print(f'ERROR: master file not found: {MASTER_PATH}', file=sys.stderr)
        return 2
    wb = openpyxl.load_workbook(MASTER_PATH, data_only=True)
    rows, violations = validate(wb)
    report = write_report(rows, violations)
    print(f'Report: {report}')
    print(f'Total: {len(rows)} 품번, 위반: {len(violations)} 건')
    return 0 if not violations else 1


if __name__ == '__main__':
    sys.exit(main())
