#!/usr/bin/env python3
"""
analyze_spec_distribution.py — TK-99-2-2 규격<7 품번 식별 + 분포 통계.

분석 항목:
    1. 압출 마스터 B열(규격) 전 품번 분포
    2. 규격<7 품번 리스트 (BR-V17 영향)
    3. 사양별 cross-tab (성형 마스터 사양 vs 압출 규격 — 일관성)

출력:
    tools/master_validation/reports/spec_distribution_<YYYY-MM-DD>.md
"""

from __future__ import annotations

import io
import sys
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EX_MASTER = Path('Phase 1/2.Raw Materials/Extrusion/압출공정_제약조건.xlsx')
VC_MASTER = Path('Phase 1/2.Raw Materials/Vulcanization/성형공정_제약조건.xlsx')
REPORT_DIR = Path('tools/master_validation/reports')


def load_ex() -> list[dict]:
    wb = openpyxl.load_workbook(EX_MASTER, data_only=True)
    ws = wb['Sheet1']
    rows = []
    for r in range(2, ws.max_row + 1):
        hose = ws.cell(r, 1).value
        if not hose:
            break
        rows.append({
            'hose': str(hose).strip(),
            'spec': ws.cell(r, 2).value,
            'inner': ws.cell(r, 3).value,
            'thick': ws.cell(r, 4).value,
            'composite': ws.cell(r, 8).value,
        })
    return rows


def load_vc_spec() -> dict[str, object]:
    """VC 마스터의 사양(B열) 매핑."""
    wb = openpyxl.load_workbook(VC_MASTER, data_only=True)
    ws = wb['Sheet1 (2)']
    result = {}
    for r in range(3, ws.max_row + 1):
        hose = ws.cell(r, 1).value
        if not hose:
            break
        result[str(hose).strip()] = ws.cell(r, 2).value
    return result


def write_report(ex_rows: list[dict], vc_spec: dict) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    out = REPORT_DIR / f'spec_distribution_{today}.md'

    spec_counter = Counter(r['spec'] for r in ex_rows)
    impact = [r for r in ex_rows if isinstance(r['spec'], (int, float)) and r['spec'] < 7]
    total = len(ex_rows)
    impact_count = len(impact)

    with open(out, 'w', encoding='utf-8') as f:
        f.write(f'# 압출 마스터 규격 분포 + BR-V17 영향 분석 ({today})\n\n')
        f.write('**참고**: BR-V17 (PDD-02 §9) — 규격<7 hose_id 는 1대 가류기당 동시 앵글 점유 ≤ 4.\n\n')

        f.write('## 1. 요약\n\n')
        f.write(f'- 총 품번: **{total}**\n')
        f.write(f'- 규격<7 품번 (BR-V17 영향): **{impact_count}** ({impact_count/total*100:.1f}%)\n')
        f.write(f'- 고유 규격: {len(spec_counter)} 종\n\n')

        f.write('## 2. 규격 분포\n\n')
        f.write('| 규격 | 품번 수 | 비율 | BR-V17 |\n|---:|---:|---:|:---:|\n')
        for s, c in sorted(spec_counter.items(), key=lambda x: (x[0] if isinstance(x[0], (int, float)) else 999)):
            mark = '✓ 영향' if isinstance(s, (int, float)) and s < 7 else ''
            f.write(f'| {s} | {c} | {c/total*100:.1f}% | {mark} |\n')
        f.write(f'| **합계** | **{total}** | 100.0% | |\n\n')

        f.write('## 3. 규격<7 품번 리스트 (BR-V17 영향)\n\n')
        f.write('| Hose | 규격 | 내경 | 두께 | 합금형 | VC 사양 |\n|---|---:|---:|---:|---:|---:|\n')
        for r in impact:
            vc_s = vc_spec.get(r['hose'], '(VC 미발견)')
            f.write(f'| {r["hose"]} | {r["spec"]} | {r["inner"]} | {r["thick"]} | {r["composite"]} | {vc_s} |\n')
        f.write('\n')

        # VC vs EX 사양 일관성 검사
        f.write('## 4. 성형 사양(VC B열) vs 압출 규격(EX B열) 일관성\n\n')
        f.write('성형 마스터의 "사양"이 압출 마스터의 "규격"과 일치해야 정상 (동일 측정 단위, 내경).\n\n')
        mismatches = []
        for r in ex_rows:
            vc_s = vc_spec.get(r['hose'])
            if vc_s is not None and vc_s != r['spec']:
                mismatches.append((r['hose'], vc_s, r['spec']))
        if mismatches:
            f.write('| Hose | VC 사양 | EX 규격 |\n|---|---:|---:|\n')
            for h, vs, es in mismatches:
                f.write(f'| ⚠️ {h} | {vs} | {es} |\n')
            f.write(f'\n**불일치 {len(mismatches)}건** — 마스터 정정 검토 필요.\n\n')
        else:
            f.write('✅ **모든 품번 VC 사양 = EX 규격** (일관성 확보)\n\n')

        # 누락 분석
        f.write('## 5. 마스터 cross-coverage\n\n')
        ex_hoses = {r['hose'] for r in ex_rows}
        vc_hoses = set(vc_spec.keys())
        only_vc = vc_hoses - ex_hoses
        only_ex = ex_hoses - vc_hoses
        f.write(f'- VC ∩ EX (둘 다 등재): {len(vc_hoses & ex_hoses)} 품번\n')
        f.write(f'- VC only (압출 미등재): **{len(only_vc)}** 품번\n')
        f.write(f'- EX only (성형 미등재): **{len(only_ex)}** 품번\n\n')
        if only_vc:
            f.write('### VC only — 압출 마스터에 없는 품번\n\n')
            for h in sorted(only_vc):
                f.write(f'- ⚠️ `{h}` (VC 사양 = {vc_spec[h]})\n')
            f.write('\n')
        if only_ex:
            f.write('### EX only — 성형 마스터에 없는 품번\n\n')
            for h in sorted(only_ex):
                f.write(f'- ⚠️ `{h}`\n')
            f.write('\n')

        f.write('\n---\n')
        f.write('## dual-review (BR-X05)\n\n')
        f.write('| 역할 | 이름 | 사인오프 일자 |\n|---|---|---|\n')
        f.write('| P1 생산계획 주임 | (서명) | (일자) |\n')
        f.write('| STK-08 IT lead | (서명) | (일자) |\n')

    return out


def main() -> int:
    if not EX_MASTER.exists() or not VC_MASTER.exists():
        print(f'ERROR: master file missing', file=sys.stderr)
        return 2
    ex_rows = load_ex()
    vc_spec = load_vc_spec()
    report = write_report(ex_rows, vc_spec)
    impact_count = sum(1 for r in ex_rows if isinstance(r['spec'], (int, float)) and r['spec'] < 7)
    print(f'Report: {report}')
    print(f'총 {len(ex_rows)} 품번 · 규격<7 영향: {impact_count} 건')
    return 0


if __name__ == '__main__':
    sys.exit(main())
