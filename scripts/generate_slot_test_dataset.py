#!/usr/bin/env python3
"""
generate_slot_test_dataset.py — TK-04-1-4 DS-VC-CONSTRAINT-47 합성기.

REF-09 (성형공정_제약조건.xlsx 47품번) → 두 파일 생성:

  1. master_seed.sql              — INSERT 47 rows into master.vc_constraint
  2. slot_violation_cases.json    — 100 (품번, 슬롯) 케이스 + expected_eligible

데이터셋 위치:
    backend/master/src/test/resources/datasets/DS-VC-CONSTRAINT-47/

REF-09 컬럼 매핑 (실측):
    col 1  HOSE          품번
    col 3  몰드수        mold_qty
    col 4  합금형        composite_count (1·2·3·6)
    col 5  LP 앵글수     lp_angle_qty
    col 6  LP 몰드/앵글  lp_molds_per_angle
    col 7  LP TOP        lp_slot_top    (o/x)
    col 8  LP UPMID      lp_slot_upmid
    col 9  LP LOWMID     lp_slot_lowmid
    col 10 LP BOT        lp_slot_bot
    col 13 IC 앵글수     ic_angle_qty
    col 14 IC 몰드/앵글  ic_molds_per_angle
    col 15 IC TOP        ic_slot_top
    col 16 IC MID        ic_slot_mid
    col 17 IC BOT        ic_slot_bot
"""

from __future__ import annotations

import io
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from random import Random

import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)

ROOT = Path(__file__).resolve().parent.parent
REF_09 = ROOT / 'Phase 1' / '2.Raw Materials' / 'Vulcanization' / '성형공정_제약조건.xlsx'
OUT_DIR = ROOT / 'backend' / 'app' / 'src' / 'test' / 'resources' / 'datasets' / 'DS-VC-CONSTRAINT-47'

# REF-09 컬럼 매핑 (Row 3 부터 데이터)
COL_HOSE = 1
COL_MOLD_QTY = 3
COL_COMPOSITE_COUNT = 4
COL_LP_ANGLE_QTY = 5
COL_LP_MOLDS_PER_ANGLE = 6
COL_LP_TOP, COL_LP_UPMID, COL_LP_LOWMID, COL_LP_BOT = 7, 8, 9, 10
COL_IC_ANGLE_QTY = 13
COL_IC_MOLDS_PER_ANGLE = 14
COL_IC_TOP, COL_IC_MID, COL_IC_BOT = 15, 16, 17

DATA_START_ROW = 3
SHEET = 'Sheet1 (2)'


def cell_bool(value) -> bool:
    """REF-09 셀 → boolean — 'o' (대소문자 + trim) = True, 그 외 = False."""
    if value is None:
        return False
    s = str(value).strip().lower()
    return s == 'o'


def cell_int(value) -> int | None:
    """셀 → int (소수점 제거). None / 빈 / 비숫자 → None."""
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def sql_int_or_null(v: int | None) -> str:
    return 'NULL' if v is None else str(v)


def sql_bool(b: bool) -> str:
    return 'true' if b else 'false'


def main() -> int:
    if not REF_09.exists():
        print(f"REF-09 미존재: {REF_09}", file=sys.stderr)
        return 1

    print(f"REF-09 load: {REF_09}")
    wb = openpyxl.load_workbook(REF_09, data_only=True)
    ws = wb[SHEET]

    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        hose = ws.cell(r, COL_HOSE).value
        if not hose:
            continue
        composite = cell_int(ws.cell(r, COL_COMPOSITE_COUNT).value)
        # composite_count 는 CHECK (1·2·3·6) — 그 외 fallback 1
        if composite not in (1, 2, 3, 6):
            composite = 1
        rows.append({
            'hose_id': str(hose).strip(),
            'mold_qty': cell_int(ws.cell(r, COL_MOLD_QTY).value) or 0,
            'composite_count': composite,
            'lp_molds_per_angle': cell_int(ws.cell(r, COL_LP_MOLDS_PER_ANGLE).value),
            'lp_angle_qty': cell_int(ws.cell(r, COL_LP_ANGLE_QTY).value),
            'lp_slot_top': cell_bool(ws.cell(r, COL_LP_TOP).value),
            'lp_slot_upmid': cell_bool(ws.cell(r, COL_LP_UPMID).value),
            'lp_slot_lowmid': cell_bool(ws.cell(r, COL_LP_LOWMID).value),
            'lp_slot_bot': cell_bool(ws.cell(r, COL_LP_BOT).value),
            'ic_molds_per_angle': cell_int(ws.cell(r, COL_IC_MOLDS_PER_ANGLE).value),
            'ic_angle_qty': cell_int(ws.cell(r, COL_IC_ANGLE_QTY).value),
            'ic_slot_top': cell_bool(ws.cell(r, COL_IC_TOP).value),
            'ic_slot_mid': cell_bool(ws.cell(r, COL_IC_MID).value),
            'ic_slot_bot': cell_bool(ws.cell(r, COL_IC_BOT).value),
        })

    print(f"Loaded {len(rows)} hose_id rows")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ----- master_seed.sql -----
    seed_path = OUT_DIR / 'master_seed.sql'
    now_iso = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S+00')
    lines = [
        "-- =============================================================================",
        f"-- DS-VC-CONSTRAINT-47 master_seed.sql — TK-04-1-4 ({len(rows)} hose_ids from REF-09)",
        "-- =============================================================================",
        "-- 자동 생성 — 수정 금지. 재생성: scripts/generate_slot_test_dataset.py",
        "",
        "DELETE FROM master.vc_constraint;",
        "",
        "INSERT INTO master.vc_constraint",
        "  (hose_id, mold_qty, composite_count,",
        "   lp_molds_per_angle, lp_angle_qty, lp_slot_top, lp_slot_upmid, lp_slot_lowmid, lp_slot_bot,",
        "   ic_molds_per_angle, ic_angle_qty, ic_slot_top, ic_slot_mid, ic_slot_bot,",
        "   updated_at, updated_by) VALUES",
    ]
    value_lines = []
    for row in rows:
        value_lines.append(
            f"  ('{row['hose_id']}', {row['mold_qty']}, {row['composite_count']}, "
            f"{sql_int_or_null(row['lp_molds_per_angle'])}, {sql_int_or_null(row['lp_angle_qty'])}, "
            f"{sql_bool(row['lp_slot_top'])}, {sql_bool(row['lp_slot_upmid'])}, "
            f"{sql_bool(row['lp_slot_lowmid'])}, {sql_bool(row['lp_slot_bot'])}, "
            f"{sql_int_or_null(row['ic_molds_per_angle'])}, {sql_int_or_null(row['ic_angle_qty'])}, "
            f"{sql_bool(row['ic_slot_top'])}, {sql_bool(row['ic_slot_mid'])}, "
            f"{sql_bool(row['ic_slot_bot'])}, "
            f"TIMESTAMPTZ '{now_iso}', 'DS-VC-CONSTRAINT-47')"
        )
    lines.append(',\n'.join(value_lines) + ';')
    lines.append("")
    seed_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f"Wrote {seed_path} ({seed_path.stat().st_size} bytes)")

    # ----- slot_violation_cases.json — 100 cases -----
    SLOT_COL_MAP = {
        'LP_TOP': 'lp_slot_top', 'LP_UPMID': 'lp_slot_upmid',
        'LP_LOWMID': 'lp_slot_lowmid', 'LP_BOT': 'lp_slot_bot',
        'IC_TOP': 'ic_slot_top', 'IC_MID': 'ic_slot_mid', 'IC_BOT': 'ic_slot_bot',
    }

    all_cells = []
    for row in rows:
        for slot_name, field in SLOT_COL_MAP.items():
            all_cells.append({
                'hose_id': row['hose_id'],
                'slot_position': slot_name,
                'expected_eligible': row[field],
            })

    eligible = [c for c in all_cells if c['expected_eligible']]
    ineligible = [c for c in all_cells if not c['expected_eligible']]
    rnd = Random(42)
    rnd.shuffle(eligible)
    rnd.shuffle(ineligible)

    # 50 적합 + 50 비적합 = 100. 부족 시 가능한 만큼.
    sampled = eligible[:50] + ineligible[:50]
    rnd.shuffle(sampled)
    cases = [{'case_id': f'SV-{i+1:03d}', **c} for i, c in enumerate(sampled)]

    cases_path = OUT_DIR / 'slot_violation_cases.json'
    cases_path.write_text(json.dumps(cases, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f"Wrote {cases_path} — {len(cases)} cases (eligible={sum(1 for c in cases if c['expected_eligible'])}, ineligible={sum(1 for c in cases if not c['expected_eligible'])})")

    # ----- unschedulable_expected.json — REF-09 zero-slot 품번 -----
    unsched = [
        row['hose_id'] for row in rows
        if not any(row[f] for f in SLOT_COL_MAP.values())
    ]
    unsched_path = OUT_DIR / 'unschedulable_expected.json'
    unsched_path.write_text(json.dumps(sorted(unsched), indent=2, ensure_ascii=False), encoding='utf-8')
    print(f"Wrote {unsched_path} — {len(unsched)} unschedulable hose_ids: {unsched}")

    return 0


if __name__ == '__main__':
    sys.exit(main())
