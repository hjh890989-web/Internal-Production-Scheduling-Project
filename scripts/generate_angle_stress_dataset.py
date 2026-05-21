#!/usr/bin/env python3
"""
generate_angle_stress_dataset.py — TK-05-2-3 DS-ANGLE-STRESS-1000 합성기.

REF-09 47품번의 lp_angle_qty / ic_angle_qty 를 입력으로 1000 stress 시나리오 생성.

각 trial:
    - 무작위 품번 선택
    - 무작위 머신 (LP-01~04 / IC-01)
    - 무작위 회전 (1~18)
    - 슬롯 점유 수 (50% 위반, 50% 정상)
    - expected_violation 정답 동시 생성 (정답 검증 자동화)

출력:
    backend/app/src/test/resources/datasets/DS-ANGLE-STRESS-1000/stress_scenarios.json

Random seed = 42 (결정적 재현).
"""

from __future__ import annotations

import io
import json
import sys
from pathlib import Path
from random import Random

import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)

ROOT = Path(__file__).resolve().parent.parent
REF_09 = ROOT / 'Phase 1' / '2.Raw Materials' / 'Vulcanization' / '성형공정_제약조건.xlsx'
OUT_DIR = ROOT / 'backend' / 'app' / 'src' / 'test' / 'resources' / 'datasets' / 'DS-ANGLE-STRESS-1000'

# REF-09 컬럼 매핑 (generate_slot_test_dataset.py 와 동일)
COL_HOSE = 1
COL_LP_ANGLE_QTY = 5
COL_IC_ANGLE_QTY = 13

DATA_START_ROW = 3
SHEET = 'Sheet1 (2)'
NUM_TRIALS = 1000
SEED = 42

LP_MACHINES = ['LP-01', 'LP-02', 'LP-03', 'LP-04']
IC_MACHINES = ['IC-01']


def cell_int(value) -> int:
    if value is None or value == '':
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def main() -> int:
    if not REF_09.exists():
        print(f"REF-09 미존재: {REF_09}", file=sys.stderr)
        return 1

    print(f"REF-09 load: {REF_09}")
    wb = openpyxl.load_workbook(REF_09, data_only=True)
    ws = wb[SHEET]

    products = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        hose = ws.cell(r, COL_HOSE).value
        if not hose:
            continue
        lp_angle = cell_int(ws.cell(r, COL_LP_ANGLE_QTY).value)
        ic_angle = cell_int(ws.cell(r, COL_IC_ANGLE_QTY).value)
        products.append({
            'hose_id': str(hose).strip(),
            'lp_angle_qty': lp_angle,
            'ic_angle_qty': ic_angle,
        })
    print(f"Loaded {len(products)} REF-09 products")

    rnd = Random(SEED)
    scenarios = []
    expected_violations = 0
    for trial in range(NUM_TRIALS):
        p = rnd.choice(products)
        # 머신 유형 — angle_qty > 0 인 쪽 우선 (그래야 의미 있는 시나리오)
        candidates = []
        if p['lp_angle_qty'] > 0:
            candidates.append(('LP', p['lp_angle_qty'], rnd.choice(LP_MACHINES)))
        if p['ic_angle_qty'] > 0:
            candidates.append(('IC', p['ic_angle_qty'], rnd.choice(IC_MACHINES)))
        if not candidates:
            # 둘 다 0 → 위반 의미 X — 슬롯 1개 (capa 0 초과)
            machine_type = rnd.choice(['LP', 'IC'])
            machine_id = rnd.choice(LP_MACHINES) if machine_type == 'LP' else 'IC-01'
            allowed = 0
        else:
            machine_type, allowed, machine_id = rnd.choice(candidates)

        # 50% 위반, 50% 정상 — n_slots 는 max_slots (LP=8, IC=6) 한도 (물리 슬롯 한도)
        max_slots = 8 if machine_type == 'LP' else 6
        if rnd.random() < 0.5:
            # 위반 시도 — allowed+1 ~ max_slots
            min_violate = allowed + 1
            if min_violate > max_slots:
                # 위반 불가 (allowed >= max_slots) — 정상 케이스로 강등
                n_slots = rnd.randint(1, max_slots)
                expected = False
            else:
                n_slots = rnd.randint(min_violate, max_slots)
                expected = True
        else:
            # 정상 — 1 ~ min(allowed, max_slots)
            cap = min(allowed, max_slots) if allowed > 0 else 0
            if cap <= 0:
                n_slots = 0
                expected = False
            else:
                n_slots = rnd.randint(1, cap)
                expected = False

        rotation_no = rnd.randint(1, 18)
        slots = []
        for slot_pos in range(1, n_slots + 1):
            slots.append({
                'machine_id': machine_id,
                'machine_type': machine_type,
                'rotation_no': rotation_no,
                'slot_position': slot_pos,
            })

        scenarios.append({
            'trial': trial,
            'hose_id': p['hose_id'],
            'machine_type': machine_type,
            'machine_id': machine_id,
            'rotation_no': rotation_no,
            'allowed_angles': allowed,
            'slot_count': n_slots,
            'slots': slots,
            'expected_violation': expected,
        })
        if expected:
            expected_violations += 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / 'stress_scenarios.json'
    out_path.write_text(json.dumps(scenarios, indent=2, ensure_ascii=False), encoding='utf-8')

    print(f"Wrote {out_path} — {len(scenarios)} trials")
    print(f"  expected violations: {expected_violations}")
    print(f"  expected normal:     {NUM_TRIALS - expected_violations}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
