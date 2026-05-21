#!/usr/bin/env python3
"""
generate_vc_alloc_scenarios.py — TK-05-3-3 DS-VC-ALLOC-100 합성기.

REF-09 schedulable 46품번에서 100 시나리오 생성:
    - 각 시나리오: 1~5 hose_id + Q_required (1~30)
    - horizon: 영업일 5일 (월~금)
    - unschedulable 품번 일부 포함 (의도적 — conflict 검증)

random seed = 42 결정적 재현.
"""

from __future__ import annotations

import io
import json
import sys
from pathlib import Path
from random import Random

import openpyxl

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)

ROOT = Path(__file__).resolve().parent.parent
REF_09 = ROOT / 'Phase 1' / '2.Raw Materials' / 'Vulcanization' / '성형공정_제약조건.xlsx'
OUT_DIR = ROOT / 'backend' / 'app' / 'src' / 'test' / 'resources' / 'datasets' / 'DS-VC-ALLOC-100'

# REF-09 컬럼 매핑 (generate_slot_test_dataset.py 와 동일)
COL_HOSE = 1
COL_LP_TOP = 7
COL_LP_UPMID = 8
COL_LP_LOWMID = 9
COL_LP_BOT = 10
COL_IC_TOP = 15
COL_IC_MID = 16
COL_IC_BOT = 17

DATA_START_ROW = 3
SHEET = 'Sheet1 (2)'
NUM_SCENARIOS = 100
SEED = 42


def cell_bool(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() == 'o'


def main() -> int:
    if not REF_09.exists():
        print(f"REF-09 미존재: {REF_09}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(REF_09, data_only=True)
    ws = wb[SHEET]

    schedulable = []
    unschedulable = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        hose = ws.cell(r, COL_HOSE).value
        if not hose:
            continue
        slot_ox = [
            cell_bool(ws.cell(r, c).value)
            for c in (COL_LP_TOP, COL_LP_UPMID, COL_LP_LOWMID, COL_LP_BOT,
                      COL_IC_TOP, COL_IC_MID, COL_IC_BOT)
        ]
        if any(slot_ox):
            schedulable.append(str(hose).strip())
        else:
            unschedulable.append(str(hose).strip())
    print(f"REF-09 — schedulable={len(schedulable)} unschedulable={len(unschedulable)}")

    rnd = Random(SEED)
    scenarios = []
    for i in range(NUM_SCENARIOS):
        # 시나리오당 1~5 hose, 일부 (10%) 는 unschedulable 포함
        n_hoses = rnd.randint(1, 5)
        include_unsched = rnd.random() < 0.10 and unschedulable

        chosen = rnd.sample(schedulable, min(n_hoses, len(schedulable)))
        if include_unsched:
            chosen.append(rnd.choice(unschedulable))

        scenario = {
            'scenario_id': f'ALLOC-{i+1:03d}',
            'description': f'{len(chosen)} hose, unschedulable_included={include_unsched}',
            'hoses': []
        }
        for hose in chosen:
            scenario['hoses'].append({
                'hose_id': hose,
                'q_required': rnd.randint(1, 30),
                'expected_unschedulable': hose in unschedulable,
            })
        scenarios.append(scenario)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / 'scenarios.json'
    out_path.write_text(json.dumps(scenarios, indent=2, ensure_ascii=False), encoding='utf-8')

    total_hoses = sum(len(s['hoses']) for s in scenarios)
    total_q = sum(h['q_required'] for s in scenarios for h in s['hoses'])
    unsched_count = sum(1 for s in scenarios for h in s['hoses'] if h['expected_unschedulable'])
    print(f"Wrote {out_path} — {len(scenarios)} scenarios, {total_hoses} hose entries, "
          f"total Q_required={total_q}, unschedulable entries={unsched_count}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
