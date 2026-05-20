#!/usr/bin/env python3
"""
generate_test_workbooks.py — TK-01-1-4 DS-ORDER-3X 합성 데이터셋 생성기

30 합성 .xlsx 워크북 (각 SourceType 7~8건 변형):
    MONTHLY_FORECAST  7건  monthly/monthly_0.xlsx ~ monthly_6.xlsx
    WEEKLY_PLAN       7건  weekly/weekly_0.xlsx  ~ weekly_6.xlsx
    CONFIRMED_ORDER   8건  confirmed/confirmed_0.xlsx ~ confirmed_7.xlsx
    KD_ORDER          8건  kd/kd_0.xlsx          ~ kd_7.xlsx

각 변형은:
    - 헤더 어휘 변형 (한글·영문 혼합)
    - 시트명 변형
    - 파일명 변형
    - 50~200 row 합성 데이터 (랜덤이지만 seed=42 → 재현 가능)

출력:
    backend/order/src/test/resources/workbooks/DS-ORDER-3X/
"""

from __future__ import annotations

import io
import random
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

random.seed(42)     # 재현 가능 (Flaky 회피)

OUT_DIR = Path(__file__).resolve().parent.parent / \
    "backend/order/src/test/resources/workbooks/DS-ORDER-3X"

# ---------------------------------------------------------------------------
# SourceType 별 변형 정의 (filename, sheetname, headers)
# ---------------------------------------------------------------------------

# 각 SourceType 별 워크북: 헤더는 hose_id·qty·delivery_date 3 필수 + (선택) 비고 컬럼.
# 분류기·매핑기 양쪽에서 다양한 어휘 변형을 검증하도록 SourceType 키워드 1개 이상 포함.

MONTHLY = [
    ("2026년 1월 월별 예상 발주.xlsx", "월간 예상",   ["품번", "수량", "예상 납기", "Σ월간"]),
    ("2026년 2월 월별 발주 예측.xlsx", "월별 예상",   ["품번", "수량", "납기", "비고"]),
    ("monthly_forecast_jan.xlsx",      "Forecast",    ["Part", "Qty", "Date"]),
    ("FORECAST_FEB_2026.xlsx",         "FORECAST",    ["Hose", "Quantity", "DueDate"]),
    ("월간 발주 예상_3월.xlsx",        "월간 예상",   ["품번", "예상 수량", "납기"]),
    ("예상 발주 4월.xlsx",             "월별 예상",   ["품번", "수량", "납품일"]),
    ("MONTHLY-FORECAST-MAY.xlsx",      "FORECAST",    ["Item", "Qty", "Date"]),
]

WEEKLY = [
    ("실리콘 02월 1주차 주간 계획.xlsx", "주간 계획", ["품번", "수량", "납기", "주차"]),
    ("주간 발주 2026-W05.xlsx",         "주간 발주", ["품번", "수량", "납기"]),
    ("weekly_plan_w05.xlsx",            "WEEKLY",    ["Hose", "Qty", "Date"]),
    ("WEEKLY_PLAN_W06.xlsx",            "WEEKLY",    ["Part", "Qty", "DueDate"]),
    ("주별 발주 W07.xlsx",              "주별 계획", ["품번", "수량", "납기"]),
    ("주차별 계획 W08.xlsx",            "주간 계획", ["품번", "수량", "납기", "비고"]),
    ("weekly-w09.xlsx",                 "Weekly",    ["Hose", "Qty", "Date"]),
]

CONFIRMED = [
    ("2026년 1월 확정 발주.xlsx",       "확정 발주",  ["품번", "확정 수량", "납기"]),
    ("2026 발주 확정.xlsx",             "발주 확정",  ["품번", "수량", "납기일"]),
    ("CONFIRMED_ORDER_JAN.xlsx",        "CONFIRMED",  ["Part", "Qty", "Date"]),
    ("confirmed_feb_2026.xlsx",         "Confirmed",  ["Hose", "Quantity", "Delivery"]),
    ("정식 발주 1주.xlsx",              "정식 발주",  ["품번", "수량", "납기"]),
    ("발주 확정_3월.xlsx",              "확정",       ["품번", "수량", "납기"]),
    ("FINAL_ORDER_W10.xlsx",            "CONFIRMED",  ["Item", "Qty", "DueDate"]),
    ("확정_W4_2026.xlsx",               "확정 발주",  ["품번", "수량", "납기"]),
]

KD = [
    ("저압 이중관 KD 발주및 납품현황 26년01월.xlsx", "KD 발주", ["품번", "KD 수량", "납품"]),
    ("KD 발주 2026-02.xlsx",                          "KD",      ["품번", "KD 수량", "납기", "재고"]),
    ("knockdown_jan.xlsx",                            "KD",      ["Part", "KD Qty", "Date"]),
    ("KD-Order-W05.xlsx",                             "KD 발주", ["품번", "수량", "납기"]),
    ("Knock-Down_2026.xlsx",                          "KnockDown", ["Hose", "Qty", "Date"]),
    ("이중관 KD 현황.xlsx",                           "KD 현황", ["품번", "수량", "납품"]),
    ("KD_INVENTORY_W08.xlsx",                         "KD",      ["Part", "Qty", "Date"]),
    ("KD_FEB_W2.xlsx",                                "KD",      ["품번", "수량", "납기"]),
]

DATASETS = [
    ("monthly", MONTHLY),
    ("weekly", WEEKLY),
    ("confirmed", CONFIRMED),
    ("kd", KD),
]

# 합성 품번 (실제 46품번 중 안전한 일부 차용 — TK-99-1-1 검증된 데이터)
HOSE_IDS = [
    "29673-2F900", "29673-2F910", "29673-2R060", "29696-2U000",
    "28421-2M100", "28422-2M100", "25450-P7200", "25451-P7200",
    "A6722030900", "A6722031002", "28415-08400", "28674-L5500",
]


def generate_workbook(out_path: Path, sheet_name: str, headers: list[str]) -> None:
    """단일 워크북 생성."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 헤더 row 1
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # 데이터 50~200 row
    n_rows = random.randint(50, 200)
    for r in range(2, n_rows + 2):
        ws.cell(row=r, column=1, value=random.choice(HOSE_IDS))
        ws.cell(row=r, column=2, value=random.randint(10, 1000))
        if len(headers) >= 3:
            ws.cell(row=r, column=3, value=f"2026-02-{(r % 28) + 1:02d}")
        for col in range(4, len(headers) + 1):
            ws.cell(row=r, column=col, value=random.randint(0, 500))

    wb.save(out_path)


def main() -> int:
    total = 0
    for subdir, dataset in DATASETS:
        for fname, sheet_name, headers in dataset:
            out_path = OUT_DIR / subdir / fname
            generate_workbook(out_path, sheet_name, headers)
            total += 1
            print(f"  ✓ {subdir}/{fname}  ({sheet_name})")
    print(f"\nDONE — {total} workbooks under {OUT_DIR}")
    print("(monthly={} / weekly={} / confirmed={} / kd={})".format(
        len(MONTHLY), len(WEEKLY), len(CONFIRMED), len(KD)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
